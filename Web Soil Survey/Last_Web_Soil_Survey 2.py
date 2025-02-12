from selenium import webdriver
# Basic libraries that define settings for webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
# Basic libraries that interacts with webpages
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

import time
# Library to handle and manipulate excel files
from openpyxl import Workbook, load_workbook

# Path to the ChromeDriver executable
CHROMEDRIVER_PATH = r"D:\Dev\Python\Selenium\\133.0.6943.54\chromedriver-win64\chromedriver.exe"

# Chrome options
chrome_options = Options()
chrome_options.add_argument('--no-sandbox')
chrome_options.add_argument('--disable-dev-shm-usage')
chrome_options.add_argument('--disable-gpu')
chrome_options.add_experimental_option('excludeSwitches', ['enable-logging'])
chrome_options.add_argument('--start-maximized')

# Initialize WebDriver
service = Service(CHROMEDRIVER_PATH)
driver = webdriver.Chrome(service=service, options=chrome_options)

# Defining default variables
default_wait_time = 5

# Generic Function to substitue all clicking functions
def locate_and_click_element_by_xpath(driver, xpath_selector, element_name):
    """
    Locates the button of the element by xpath and clicks it.
    
    Args:
        driver: The Selenium WebDriver instance.
        xpath_selector: The Xpath selector of the element.
        element_name: The name of the element to click.
    """
    try:
        # Wait for the element to be present and clickable
        expand_button = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, xpath_selector))
        )
        
        # Click the button
        expand_button.click()
        print(f"{element_name} clicked successfully.")
    except Exception as e:
        print(f"Failed to locate or click {element_name}: {e}")

# Generic Function to substitue all filling functions
def locate_and_fill_field_input_by_xpath(driver, input_value, xpath_selector, input_name):
    """
    Fills input text area with the provided input_value using XPath.

    Args:
        driver: The Selenium WebDriver instance.
        input_value (str): The value to fill into the textarea.
        xpath_location: The Xpath selector of the element.
        input_name: The input name of the element.
    """
    try:
        # Wait for the textarea to be present using XPath
        element_textarea = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, xpath_selector))
        )
        
        # Clear any existing text (optional)
        element_textarea.clear()
        
        # Fill the textarea with the test_address
        element_textarea.send_keys(input_value)
        print(f"{input_name} filled with: {input_value}")
    except Exception as e:
        print(f"Failed to locate or fill the {input_name}: {e}")

def extract_table_data_one(driver):
    # Define the XPath for the table
    table_xpath = "//table[contains(@class, 'data') and contains(@class, 'first') and contains(@class, 'last')]"
    
    # Define the XPath for all rows inside the tbody section of the table
    rows_xpath = f"{table_xpath}/tbody/tr"
    
    # Initialize a list to store extracted data
    table_data = []
    
    # Locate all table rows within the tbody
    rows = driver.find_elements(By.XPATH, rows_xpath)
    
    # Loop through each row to extract data
    for row in rows:
        # Find all the columns (td elements) in the current row
        cells = row.find_elements(By.XPATH, ".//td")
        
        # Extract text from each cell and strip any extra whitespace
        row_data = [cell.text.strip() for cell in cells]
        
        # Convert row data into a dictionary using column labels as keys
        if len(row_data) == 4:  # Ensure we have the correct number of columns
            row_dict = {
                "Map Unit Symbol": row_data[0],
                "Map Unit Name": row_data[1],
                "Acres in AOI": row_data[2],
                "Percent of AOI": row_data[3]
            }
            table_data.append(row_dict)
    
    return table_data

def extract_table_data_by_xpath(driver, xpath_selector):
    # Define the XPath for the table
    table_xpath = xpath_selector
    
    # Define the XPath for all rows inside the tbody section of the table
    rows_xpath = f"{table_xpath}/tbody/tr" 
    
    # Initialize a list to store extracted data
    table_data = []
    
    # Locate all table rows within the tbody
    rows = driver.find_elements(By.XPATH, rows_xpath)
    
    # Loop through each row to extract data
    for row in rows:
        # Find all the columns (td elements) in the current row
        cells = row.find_elements(By.XPATH, ".//td")
        
        # Extract text from each cell and strip any extra whitespace
        row_data = [cell.text.strip() for cell in cells]
        
        # Convert row data into a dictionary using column labels as keys
        if len(row_data) == 5:  # Ensure we have the correct number of columns
            row_dict = {
                "Map Unit Symbol": row_data[0],
                "Map Unit Name": row_data[1],
                "Rating": row_data[2],
                "Acres in AOI": row_data[3],
                "Percent of AOI": row_data[4]
            }
            table_data.append(row_dict)
    
    return table_data

'''
def find_highest_percent_aoi(data_list):
    # Convert 'Percent of AOI' to numeric and find the max
    highest = max(data_list, key=lambda x: float(x['Percent of AOI'].strip('%')))
    return highest["Map Unit Symbol"], highest["Map Unit Name"], highest["Rating"]
'''

def find_highest_percent_aoi(data_list):
    if not data_list:
        print("Warning: Input data list is empty.")
        return []  # Return an empty list if the input is empty

    try:
        # Convert 'Percent of AOI' to numeric and find the max
        highest = max(data_list, key=lambda x: float(x['Percent of AOI'].strip('%')))
        return [{
            "Map Unit Symbol": highest["Map Unit Symbol"],
            "Map Unit Name": highest["Map Unit Name"],
            "Rating": highest["Rating"]
        }]
    except (KeyError, ValueError) as e:
        print(f"Error occurred: {e}")
        return []  # Return an empty list if there's an error

def fill_address_and_click_view (driver, address):
    #locating and clicking expand address button
    xpath_expand_address_button = "//span[@class='right']//span[@class='unfold' and @role='button' and .//span[@class='sr-only' and text()='Expand Address']]"
    name_expand_address_button = "Expand Address Button"
    locate_and_click_element_by_xpath (driver, xpath_expand_address_button, name_expand_address_button)
    time.sleep(1)

    # locating and filling the text area of current address
    xpath_fill_address = "//textarea[@name='address' and @label='Address']"
    name_fill_address = "Address Text Area"
    input_fill_address = address
    locate_and_fill_field_input_by_xpath(driver, input_fill_address, xpath_fill_address, name_fill_address)
    time.sleep(1)

    # locating and clicking the view button
    xpath_view_button = "//*[@id='navigatebyaddressformid']/div[1]/button"
    name_view_button = "View Button"
    locate_and_click_element_by_xpath (driver, xpath_view_button, name_view_button)

    # pause and wait for user input (Press Enter to Continue)
    input("Please Adjust the View, Draw the AOI and Press Enter to continue...")

def extract_data_from_soil_map_tab (driver):
    # locating and clicking soil map tab (we can remove this since we have the same data in the next tab)
    xpath_soil_map_tab = "//td[@class='tabs']//td[@class='tab-inactive']/div[@id='Soil_Map']"
    name_soil_map_tab = "Soil Map Tab"
    locate_and_click_element_by_xpath (driver, xpath_soil_map_tab, name_soil_map_tab)
    time.sleep (default_wait_time)
        
    # extracting the data from the soil map
    data_in_soil_map = extract_table_data_one (driver)
    print (data_in_soil_map)
    time.sleep (default_wait_time)
    
def extract_data_from_suitabilities_and_limitations_for_use_tab (driver):
    # locating and clicking the expand land classification
    xpath_expand_land_classification_button = "//span[@class='right']//span[@class='unfold' and @role='button' and .//span[@class='sr-only' and text()='Expand Land Classifications']]"
    name_expand_land_classification_button = "Expand Land Classification Button"
    locate_and_click_element_by_xpath (driver, xpath_expand_land_classification_button, name_expand_land_classification_button)
    time.sleep (default_wait_time)
        
    # locating and clicking the hydric rating by map unit button
    xpath_hydric_rating_by_map_unit_button = "//div[@id='Land_Classifications_Hydric_Rating_by_Map_Unit_title']"
    name_hydric_rating_by_map_unit_button = "Hydric Rating By Map Unit Button"
    locate_and_click_element_by_xpath (driver, xpath_hydric_rating_by_map_unit_button, name_hydric_rating_by_map_unit_button)
    time.sleep (default_wait_time)

    # locating and clicking the view rating for hydric rating by map unit
    xpath_view_rating_for_hydric_rating_by_map_unit_button = "//button[@id='ParameterForm16ViewRating_top']"
    name_view_rating_for_hydric_rating_by_map_unit_button = "View Rating For Hydric Rating By Map Unit Button"
    locate_and_click_element_by_xpath (driver, xpath_view_rating_for_hydric_rating_by_map_unit_button, name_view_rating_for_hydric_rating_by_map_unit_button)
    time.sleep (default_wait_time)

    # extracting data from hydric rating by map unit
    xpath_hydric_rating_by_map_unit_table = "//*[@id='bymapunittablepanel_id_body']/div/table"
    data_in_hydric_rating_by_map_unit = extract_table_data_by_xpath (driver, xpath_hydric_rating_by_map_unit_table)
    print (data_in_hydric_rating_by_map_unit)
    hydric_rating_data = find_highest_percent_aoi (data_in_hydric_rating_by_map_unit)
    
    print ('hydric_rating_data ', hydric_rating_data)

    # locating and clicking the expand soil moisture class button
    xpath_expand_soil_moisture_button = "//*[@id='Land_Classifications_Soil_Moisture_Class_title']"
    name_expand_soil_moisture_button = "Soil Moisture Class Button"
    locate_and_click_element_by_xpath (driver, xpath_expand_soil_moisture_button, name_expand_soil_moisture_button)
    time.sleep (default_wait_time)

    # locating and clicking the expand soil moisture class button
    xpath_expand_soil_moisture_button = "//*[@id='ParameterForm2790ViewRating_top']"
    name_expand_soil_moisture_button = "View Rating Soil Moisture Class Button"
    locate_and_click_element_by_xpath (driver, xpath_expand_soil_moisture_button, name_expand_soil_moisture_button)
    time.sleep (default_wait_time)

    # extracting data from soil moisture class
    xpath_soil_moisture_class_table = "//*[@id='bymapunittablepanel_id_body']/div/table"
    data_in_soil_moisture_class_table = extract_table_data_by_xpath (driver, xpath_soil_moisture_class_table)
    print (data_in_soil_moisture_class_table)
    soil_moisture_class_data = find_highest_percent_aoi(data_in_soil_moisture_class_table)
    print ('soil_moisture_class_data ', soil_moisture_class_data)
    
def extract_available_water_capacity (driver):
    xpath_available_water_capacity = '//*[@id="Soil_Physical_Properties_Available_Water_Capacity_header"]/span/span'
    name_available_water_capacity = 'Available Water Capacity'
    locate_and_click_element_by_xpath (driver, xpath_available_water_capacity, name_available_water_capacity)
    time.sleep (default_wait_time)
    
    xpath_fill_top_depth = "//input[@type='text' and @name='Top_Depth' and @label='Top Depth']"
    name_fill_top_depth = 'Top Depth'
    input_value_fill_top_depth = '0'
    locate_and_fill_field_input_by_xpath(driver, input_value_fill_top_depth, xpath_fill_top_depth, name_fill_top_depth)
    time.sleep (default_wait_time)
    
    xpath_fill_bottom_depth = "//input[@type='text' and @name='Bottom_Depth' and @label='Bottom Depth']"
    name_fill_bottom_depth = 'Bottom Depth'
    input_value_fill_bottom_depth = '150'
    locate_and_fill_field_input_by_xpath(driver, input_value_fill_bottom_depth, xpath_fill_bottom_depth, name_fill_bottom_depth)
    time.sleep (default_wait_time)
    
    xpath_view_rating_button = '//*[@id="ParameterForm41ViewRating_top"]'
    name_view_rating_button = 'View Rating for Available Water Capacity'
    locate_and_click_element_by_xpath (driver, xpath_view_rating_button, name_view_rating_button)
    time.sleep (default_wait_time)
    
    # extracting data from soil physical properties, available water capacity
    xpath_available_water_capacity_table = '//*[@id="bymapunittablepanel_id_body"]/div/table'
    data_in_available_water_capacity_table = extract_table_data_by_xpath (driver, xpath_available_water_capacity_table)
    print (data_in_available_water_capacity_table)
    available_water_capacity_data = find_highest_percent_aoi(data_in_available_water_capacity_table)
    print ('available_water_capacity max percentage ', available_water_capacity_data)
    
def extract_bulk_density_one_third_bar (driver):
    xpath_bulk_density_button = '//*[@id="Soil_Physical_Properties_Bulk_Density.44._One.45.Third_Bar_header"]/span' 
    name_bulk_density_button = 'Bulk Density, One-Third Bar'
    locate_and_click_element_by_xpath (driver, xpath_bulk_density_button, name_bulk_density_button)
    time.sleep (default_wait_time)
    
    xpath_fill_top_depth = "//input[@type='text' and @name='Top_Depth' and @label='Top Depth' and contains(@class, 'numeric')]"
    name_fill_top_depth = 'Top Depth'
    input_value_fill_top_depth = '0'
    locate_and_fill_field_input_by_xpath(driver, input_value_fill_top_depth, xpath_fill_top_depth, name_fill_top_depth)
    time.sleep (default_wait_time)
    
    xpath_fill_bottom_depth = "//input[@type='text' and @name='Bottom_Depth' and @label='Bottom Depth' and contains(@class, 'numeric')]"
    name_fill_bottom_depth = 'Bottom Depth'
    input_value_fill_bottom_depth = '150'
    locate_and_fill_field_input_by_xpath(driver, input_value_fill_bottom_depth, xpath_fill_bottom_depth, name_fill_bottom_depth)
    time.sleep (default_wait_time)
    
    input("Please Adjust the View, Draw the AOI and Press Enter to continue...")
    
    xpath_view_rating_button = '//*[@id="ParameterForm65ViewRating_top"]'
    name_view_rating_button = 'View Rating for Available Water Capacity'
    locate_and_click_element_by_xpath (driver, xpath_view_rating_button, name_view_rating_button)
    time.sleep (default_wait_time)
    
    # extracting data from soil physical properties, bulk density, one-third bard
    xpath_bulk_density_table = '//*[@id="bymapunittablepanel_id_body"]/div/table'
    data_in_bulk_density_table = extract_table_data_by_xpath (driver, xpath_bulk_density_table)
    print (data_in_bulk_density_table)
    available_water_capacity_data = find_highest_percent_aoi(data_in_bulk_density_table)
    print ('available_water_capacity max percentage ', data_in_bulk_density_table)
    
def extract_liquid_limit (driver):
    xpath_liquid_limit_button = '//*[@id="Soil_Physical_Properties_Liquid_Limit_header"]/span/span'
    name_liquid_limit_button = 'Liquid Limit'
    locate_and_click_element_by_xpath (driver, xpath_liquid_limit_button, name_liquid_limit_button)
    time.sleep (default_wait_time)
    
    xpath_fill_top_depth = "//input[@name='Top_Depth']"
    name_fill_top_depth = 'Top Depth'
    input_value_fill_top_depth = '0'
    locate_and_fill_field_input_by_xpath(driver, input_value_fill_top_depth, xpath_fill_top_depth, name_fill_top_depth)
    time.sleep (default_wait_time)
    
    xpath_fill_bottom_depth = "//input[@name='Bottom_Depth']"
    name_fill_bottom_depth = 'Bottom Depth'
    input_value_fill_bottom_depth = '150'
    locate_and_fill_field_input_by_xpath(driver, input_value_fill_bottom_depth, xpath_fill_bottom_depth, name_fill_bottom_depth)
    time.sleep (default_wait_time)
    
    xpath_view_rating_button = '//*[@id="ParameterForm378ViewRating_top"]'
    name_view_rating_button = 'View Rating for Liquid Limit'
    locate_and_click_element_by_xpath (driver, xpath_view_rating_button, name_view_rating_button)
    time.sleep (default_wait_time)
    
    # extracting data from soil physical properties, bulk density, one-third bard
    xpath_liquid_limit_table = '//*[@id="bymapunittablepanel_id_body"]/div/table'
    data_in_liquid_limit_table = extract_table_data_by_xpath (driver, xpath_liquid_limit_table)
    print (data_in_liquid_limit_table)
    liquid_limit_data = find_highest_percent_aoi(data_in_liquid_limit_table)
    print ('liquid_limit max percentage ', liquid_limit_data)
    
def extract_palsticity_index (driver):
    xpath_plasticity_index_button = '//*[@id="Soil_Physical_Properties_Plasticity_Index_header"]/span/span'
    name_plasticity_index_button = 'Plasticity Index'
    locate_and_click_element_by_xpath (driver, xpath_plasticity_index_button, name_plasticity_index_button)
    time.sleep (default_wait_time)
    
    xpath_fill_top_depth = "//input[@name='Top_Depth']"
    name_fill_top_depth = 'Top Depth'
    input_value_fill_top_depth = '0'
    locate_and_fill_field_input_by_xpath(driver, input_value_fill_top_depth, xpath_fill_top_depth, name_fill_top_depth)
    time.sleep (default_wait_time)
    
    xpath_fill_bottom_depth = "//input[@name='Bottom_Depth']"
    name_fill_bottom_depth = 'Bottom Depth'
    input_value_fill_bottom_depth = '150'
    locate_and_fill_field_input_by_xpath(driver, input_value_fill_bottom_depth, xpath_fill_bottom_depth, name_fill_bottom_depth)
    time.sleep (default_wait_time)
    
    xpath_view_rating_button = '//*[@id="ParameterForm379ViewRating_top"]'
    name_view_rating_button = 'View Rating for Plasticity Index'
    locate_and_click_element_by_xpath (driver, xpath_view_rating_button, name_view_rating_button)
    time.sleep (default_wait_time)
    
    # extracting data from soil physical properties, bulk density, one-third bard
    xpath_plasticity_index_table = '//*[@id="bymapunittablepanel_id_body"]/div/table'
    data_in_plasticity_index_table = extract_table_data_by_xpath (driver, xpath_plasticity_index_table)
    print (data_in_plasticity_index_table)
    plasticity_index_data = find_highest_percent_aoi(data_in_plasticity_index_table)
    print ('plasticity_index max percentage ', plasticity_index_data)

def extract_saturated_hydraulic_conductivity (driver):
    xpath_saturated_hydraulic_conductivity = '//*[@id="Soil_Physical_Properties_Saturated_Hydraulic_Conductivity_.40.Ksat.41._header"]/span/span'
    name_saturated_hydraulic_conductivity_button = 'Saturated Hydraulic Conductivity (Ksat)'
    locate_and_click_element_by_xpath (driver, xpath_saturated_hydraulic_conductivity, name_saturated_hydraulic_conductivity_button)
    time.sleep (default_wait_time)
    
    xpath_fill_top_depth = "//input[@name='Top_Depth']"
    name_fill_top_depth = 'Top Depth'
    input_value_fill_top_depth = '0'
    locate_and_fill_field_input_by_xpath(driver, input_value_fill_top_depth, xpath_fill_top_depth, name_fill_top_depth)
    time.sleep (default_wait_time)
    
    xpath_fill_bottom_depth = "//input[@name='Bottom_Depth']"
    name_fill_bottom_depth = 'Bottom Depth'
    input_value_fill_bottom_depth = '150'
    locate_and_fill_field_input_by_xpath(driver, input_value_fill_bottom_depth, xpath_fill_bottom_depth, name_fill_bottom_depth)
    time.sleep (default_wait_time)
    
    xpath_view_rating_button = '//*[@id="ParameterForm42ViewRating_top"]'
    name_view_rating_button = 'View Rating for Saturated Hydraulic Conductivity (Ksat)'
    locate_and_click_element_by_xpath (driver, xpath_view_rating_button, name_view_rating_button)
    time.sleep (default_wait_time)
    
    # extracting data from soil physical properties, bulk density, one-third bard
    xpath_saturated_hydraulic_conductivity_table = '//*[@id="bymapunittablepanel_id_body"]/div/table'
    data_saturated_hydraulic_conductivity_table = extract_table_data_by_xpath (driver, xpath_saturated_hydraulic_conductivity_table)
    print (data_saturated_hydraulic_conductivity_table)
    saturated_hydraulic_conductivity_data = find_highest_percent_aoi(data_saturated_hydraulic_conductivity_table)
    print ('saturated_hydraulic_conductivity max percentage ', saturated_hydraulic_conductivity_data)
    
def extract_surface_texture (driver):
    xpath_of_element = '//*[@id="Soil_Physical_Properties_Surface_Texture_header"]/span/span'
    name_of_element = 'Surface Texture'
    locate_and_click_element_by_xpath (driver, xpath_of_element, name_of_element)
    
    time.sleep (default_wait_time)
    
    # skipping fill the input with values
    '''
    xpath_fill_top_depth = "//input[@name='Top_Depth']"
    name_fill_top_depth = 'Top Depth'
    input_value_fill_top_depth = '0'
    locate_and_fill_field_input_by_xpath(driver, input_value_fill_top_depth, xpath_fill_top_depth, name_fill_top_depth)
    time.sleep (default_wait_time)
    
    xpath_fill_bottom_depth = "//input[@name='Bottom_Depth']"
    name_fill_bottom_depth = 'Bottom Depth'
    input_value_fill_bottom_depth = '150'
    locate_and_fill_field_input_by_xpath(driver, input_value_fill_bottom_depth, xpath_fill_bottom_depth, name_fill_bottom_depth)
    time.sleep (default_wait_time)
    '''
    
    # clicking View Rating for Surface Texture
    xpath_button = '//*[@id="ParameterForm308ViewRating_top"]'
    name_button = 'View Rating for Surface Texture'
    locate_and_click_element_by_xpath (driver, xpath_button, name_button)
    time.sleep (default_wait_time)
    
    # extracting data from soil physical properties, bulk density, one-third bard
    xpath_table = '//*[@id="bymapunittablepanel_id_body"]/div/table'
    data_table = extract_table_data_by_xpath (driver, xpath_table)
    print (data_table)
    final_data = find_highest_percent_aoi(data_table)
    print ('surface texture max percentage ', final_data)
    
def extract_data_from_soil_properties_and_qualities_tab (driver):
    xpath_soil_properties_and_qaulities_tab = '//*[@id="Soil_Properties_and_Qualities"]'
    name_soil_properties_and_qaulities_tab = 'Soil Properties and Qualities'
    locate_and_click_element_by_xpath (driver, xpath_soil_properties_and_qaulities_tab, name_soil_properties_and_qaulities_tab)
    time.sleep (default_wait_time)
    
    # locating and clicking soil data explorer
    xpath_physical_properties_button = '//*[@id="Soil_Physical_Properties_unfold"]' 
    name_physical_properties_button = 'Soil Physical Properties'
    locate_and_click_element_by_xpath (driver, xpath_physical_properties_button, name_physical_properties_button)
    time.sleep (default_wait_time)

    extract_available_water_capacity (driver)
    input("Please Adjust the View, Draw the AOI and Press Enter to continue...")
    
    extract_bulk_density_one_third_bar (driver)
    input("Please Adjust the View, Draw the AOI and Press Enter to continue...")
    
    extract_liquid_limit (driver)
    input("Please Adjust the View, Draw the AOI and Press Enter to continue...")
    
    extract_palsticity_index (driver)
    input("Please Adjust the View, Draw the AOI and Press Enter to continue...")

    extract_saturated_hydraulic_conductivity (driver)
    input("Please Adjust the View, Draw the AOI and Press Enter to continue...")

    extract_surface_texture (driver)
    input("Please Adjust the View, Draw the AOI and Press Enter to continue...")


def extract_data_from_soil_data_explorer_tab (driver):
    # locating and clicking soil data explorer
    xpath_soil_data_explorer_tab = "//td[@class='tabs']//td[@class='tab-inactive']/div[@id='Soil_Data_Explorer']"
    name_soil_map_tab = "Soil Data Explorer"
    locate_and_click_element_by_xpath (driver, xpath_soil_data_explorer_tab, name_soil_map_tab)
    time.sleep (default_wait_time)

    #extract_data_from_suitabilities_and_limitations_for_use_tab (driver)

    extract_data_from_soil_properties_and_qualities_tab (driver)
    

# Main variables 
base_url = 'https://websoilsurvey.nrcs.usda.gov/app/WebSoilSurvey.aspx'
#test_address = '524 boston port madison ct'
addresses_file_path = 'addresses_list.xlsx'

# Start the browser automation
try:
    print(f"Processing file: {addresses_file_path}")
    # Load the workbook and sheet
    workbook = load_workbook(addresses_file_path)
    sheet = workbook.active
    
    for row_index, column in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):  # Skip header row
        current_address = column[0]  # Loading addresses from first column
        print ('Processing address: ', current_address)
        # Initialize WebDriver
        service = Service(CHROMEDRIVER_PATH)
        driver = webdriver.Chrome(service=service, options=chrome_options)
        print('Initializing browser settings ...')
        
        #time.sleep(default_wait_time)
        driver.get(base_url)
        print ('Visiting webpage')
        time.sleep(10)

        fill_address_and_click_view (driver, current_address)

        #extract_data_from_soil_map_tab (driver)
        
        extract_data_from_soil_data_explorer_tab (driver)
        #print ('map_unit_symbol :', map_unit_symbol , 'map_unit_name ', map_unit_name, 'rating', rating)



except Exception as e:
    print("Error occurred: ", e)
finally:
    driver.quit()

