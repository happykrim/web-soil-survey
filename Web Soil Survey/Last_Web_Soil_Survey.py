from selenium import webdriver
# Basic libraries that define settings for webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
# Basic libraries that interacts with webpages
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

import time
# Library to handle and manipulate excel files
from openpyxl import Workbook, load_workbook

# Path to the ChromeDriver executable
CHROMEDRIVER_PATH = r"D:\Dev\Python\Selenium\\133.0.6943.54\chromedriver-win64\chromedriver.exe"

# Chrome options
chrome_options = Options()
chrome_options.add_argument('--no-sandbox')
chrome_options.add_argument('--disable-dev-shm-usage')
chrome_options.add_argument('--disable-gpu')
chrome_options.add_experimental_option('excludeSwitches', ['enable-logging'])
chrome_options.add_argument('--start-maximized')

# Initialize WebDriver
service = Service(CHROMEDRIVER_PATH)
driver = webdriver.Chrome(service=service, options=chrome_options)

def scroll_to_bottom_of_the_page(driver, scroll_step=300, delay=0.5):
    """
    Scrolls down to the bottom of the page slowly.
    """
    last_height = driver.execute_script("return document.body.scrollHeight")
    current_position = 0

    while current_position < last_height:
        driver.execute_script(f"window.scrollTo(0, {current_position});")
        current_position += scroll_step
        time.sleep(delay)
        last_height = driver.execute_script("return document.body.scrollHeight")

    driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
    

# Generic Function to substitue all clicking functions
def locate_and_click_element_by_xpath(driver, xpath_location, element_name):
    """
    Locates the button of the element by xpath and clicks it.
    
    Args:
        driver: The Selenium WebDriver instance.
        xpath: The Xpath of the element.
    """
    try:
        # Wait for the element to be present and clickable
        expand_button = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, xpath_location))
        )
        
        # Click the button
        expand_button.click()
        print(f"{element_name} clicked successfully.")
    except Exception as e:
        print(f"Failed to locate or click {element_name}: {e}")    
        
# Function to locate expand address
def locate_and_click_expand_address_button(driver):
    """
    Locates the 'Expand Address' button and clicks it.
    
    Args:
        driver: The Selenium WebDriver instance.
    """
    try:
        # Wait for the element to be present and clickable
        expand_button = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, "//span[@class='right']//span[@class='unfold' and @role='button' and .//span[@class='sr-only' and text()='Expand Address']]"))
        )
        
        # Click the button
        expand_button.click()
        print("Expand Address button clicked successfully.")
    except Exception as e:
        print(f"Failed to locate or click the Expand Address button: {e}")

# Function to fill the address input
def fill_address_input(driver, test_address):
    """
    Fills the address textarea with the provided test_address using XPath.
    
    Args:
        driver: The Selenium WebDriver instance.
        test_address (str): The address to fill into the textarea.
    """
    try:
        # Wait for the textarea to be present using XPath
        address_textarea = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "//textarea[@name='address' and @label='Address']"))
        )
        
        # Clear any existing text (optional)
        address_textarea.clear()
        
        # Fill the textarea with the test_address
        address_textarea.send_keys(test_address)
        print(f"Address textarea filled with: {test_address}")
    except Exception as e:
        print(f"Failed to locate or fill the address textarea: {e}")
        
def locate_and_click_view_button(driver):
    """
    Locates the 'View' button and clicks it.
    
    Args:
        driver: The Selenium WebDriver instance.
    """
    try:
        # Wait for the element to be present and clickable
        expand_button = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, "//*[@id='navigatebyaddressformid']/div[1]/button"))
        )
        
        # Click the button
        expand_button.click()
        print("View button clicked successfully.")
    except Exception as e:
        print(f"Failed to locate or click the Expand Address button: {e}")
        
# Function to locate expand address
def locate_and_click_soil_map_tab(driver):
    """
    Locates the 'Soil Map' tab and clicks it.
    
    Args:
        driver: The Selenium WebDriver instance.
    """
    try:
        # Wait for the element to be present and clickable
        expand_button = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, "//td[@class='tabs']//td[@class='tab-inactive']/div[@id='Soil_Map']"))
        )
        
        # Click the button
        expand_button.click()
        print("Soil Map tab button clicked successfully.")
    except Exception as e:
        print(f"Failed to locate or click the Soil Map tab: {e}")
        
# Function to locate expand address
def locate_and_click_soil_data_explorer(driver):
    """
    Locates the 'Soil Data Explorer' tab and clicks it.
    
    Args:
        driver: The Selenium WebDriver instance.
    """
    try:
        # Wait for the element to be present and clickable
        expand_button = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, "//td[@class='tabs']//td[@class='tab-inactive']/div[@id='Soil_Data_Explorer']"))
        )
        
        # Click the button
        expand_button.click()
        print("Soil Data Explorer tab button clicked successfully.")
    except Exception as e:
        print(f"Failed to locate or click the Soil Map tab: {e}")
        
        
# Function to locate expand land classification button
def locate_and_click_expand_hydric_rating_by_map_unit_button(driver):
    """
    Locates the 'Expand Address' button and clicks it.
    
    Args:
        driver: The Selenium WebDriver instance.
    """
    try:
        # Wait for the element to be present and clickable
        expand_button = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, "//span[@class='right']//span[@class='unfold' and @role='button' and .//span[@class='sr-only' and text()='Expand Land Classifications']]"))
        )
        
        # Click the button
        expand_button.click()
        print("Expand Land Classifications button clicked successfully.")
    except Exception as e:
        print(f"Failed to locate or click the Expand Land Classifications button: {e}")
        
# Function to locate expand land classification button
def locate_and_click_expand_land_classification_button(driver):
    """
    Locates the 'Expand Address' button and clicks it.
    
    Args:
        driver: The Selenium WebDriver instance.
    """
    try:
        # Wait for the element to be present and clickable
        expand_button = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, "//span[@class='right']//span[@class='unfold' and @role='button' and .//span[@class='sr-only' and text()='Expand Land Classifications']]"))
        )
        
        # Click the button
        expand_button.click()
        print("Expand Land Classifications button clicked successfully.")
    except Exception as e:
        print(f"Failed to locate or click the Expand Land Classifications button: {e}")
        
# Function to locate expand address
def locate_and_click_hydric_rating_by_map_unit_button(driver):
    """
    Locates the 'Expand Address' button and clicks it.
    
    Args:
        driver: The Selenium WebDriver instance.
    """
    try:
        # Wait for the element to be present and clickable
        expand_button = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, "//div[@id='Land_Classifications_Hydric_Rating_by_Map_Unit_title']"))
        )
        
        # Click the button
        expand_button.click()
        print("Hydric Rating By Map Unit button clicked successfully.")
    except Exception as e:
        print(f"Failed to locate or click the Hydric Rating By Map Unit button: {e}")
        
# Function to locate expand address
def locate_and_click_view_rating_for_hydric_rating_by_map_unit_button(driver):
    """
    Locates the 'Expand Address' button and clicks it.
    
    Args:
        driver: The Selenium WebDriver instance.
    """
    try:
        # Wait for the element to be present and clickable
        expand_button = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, "//button[@id='ParameterForm16ViewRating_top']"))
        )
        
        # Click the button
        expand_button.click()
        print("View Rating for Hydric Rating button clicked successfully.")
    except Exception as e:
        print(f"Failed to locate or click the View Rating for Hydric Rating button: {e}")

def extract_table_data_one(driver):
    # Define the XPath for the table
    table_xpath = "//table[contains(@class, 'data') and contains(@class, 'first') and contains(@class, 'last')]"
    
    # Define the XPath for all rows inside the tbody section of the table
    rows_xpath = f"{table_xpath}/tbody/tr"
    
    # Initialize a list to store extracted data
    table_data = []
    
    # Locate all table rows within the tbody
    rows = driver.find_elements(By.XPATH, rows_xpath)
    
    # Loop through each row to extract data
    for row in rows:
        # Find all the columns (td elements) in the current row
        cells = row.find_elements(By.XPATH, ".//td")
        
        # Extract text from each cell and strip any extra whitespace
        row_data = [cell.text.strip() for cell in cells]
        
        # Convert row data into a dictionary using column labels as keys
        if len(row_data) == 4:  # Ensure we have the correct number of columns
            row_dict = {
                "Map Unit Symbol": row_data[0],
                "Map Unit Name": row_data[1],
                "Acres in AOI": row_data[2],
                "Percent of AOI": row_data[3]
            }
            table_data.append(row_dict)
    
    return table_data

def extract_table_data_two(driver):
    # Define the XPath for the table
    table_xpath = "//*[@id='bymapunittablepanel_id_body']/div/table"
    
    # Define the XPath for all rows inside the tbody section of the table
    rows_xpath = f"{table_xpath}/tbody/tr"
    
    # Initialize a list to store extracted data
    table_data = []
    
    # Locate all table rows within the tbody
    rows = driver.find_elements(By.XPATH, rows_xpath)
    
    # Loop through each row to extract data
    for row in rows:
        # Find all the columns (td elements) in the current row
        cells = row.find_elements(By.XPATH, ".//td")
        
        # Extract text from each cell and strip any extra whitespace
        row_data = [cell.text.strip() for cell in cells]
        
        # Convert row data into a dictionary using column labels as keys
        if len(row_data) == 5:  # Ensure we have the correct number of columns
            row_dict = {
                "Map Unit Symbol": row_data[0],
                "Map Unit Name": row_data[1],
                "Rating": row_data[2],
                "Acres in AOI": row_data[3],
                "Percent of AOI": row_data[4]
            }
            table_data.append(row_dict)
    
    return table_data

def find_highest_percent_aoi(data_list):
    # Convert 'Percent of AOI' to numeric and find the max
    highest = max(data_list, key=lambda x: float(x['Percent of AOI'].strip('%')))
    return highest["Map Unit Symbol"], highest["Map Unit Name"], highest["Rating"]

# Main variables 
base_url = 'https://websoilsurvey.nrcs.usda.gov/app/WebSoilSurvey.aspx'
test_address = '524 boston port madison ct'
addresses_file_path = 'addresses_list.xlsx'

# Start the browser automation
try:
    print(f"Processing file: {addresses_file_path}")
    # Load the workbook and sheet
    workbook = load_workbook(addresses_file_path)
    sheet = workbook.active
    
    for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):  # Skip header row
        current_address = row[0]  # Loading addresses from first column
        print ('Processing address: ', current_address)
        # Initialize WebDriver
        service = Service(CHROMEDRIVER_PATH)
        driver = webdriver.Chrome(service=service, options=chrome_options)
        print('Initializing browser settings ...')
        
        #time.sleep(5)
        driver.get(base_url)
        print ('Visiting webpage')
        time.sleep(10)
        
        #locating and clicking expand address button
        xpath_expand_address_button = "//span[@class='right']//span[@class='unfold' and @role='button' and .//span[@class='sr-only' and text()='Expand Address']]"
        locate_and_click_element_by_xpath (driver, )
        locate_and_click_expand_address_button (driver)
        time.sleep(1)
        
        fill_address_input (driver, current_address)
        time.sleep(1)
        
        locate_and_click_view_button(driver)

        # Pause and wait for user input
        input("Set View, Adjust View and Draw AOI and Press Enter to continue...")
        time.sleep (2)

        locate_and_click_soil_map_tab (driver)
        
        time.sleep (10)
        
        data_in_soil_map = extract_table_data_one (driver)
        print (data_in_soil_map)
        time.sleep (10)
        
        locate_and_click_soil_data_explorer(driver)
        time.sleep (10)
        
        locate_and_click_expand_land_classification_button(driver)
        time.sleep (5)
        
        locate_and_click_hydric_rating_by_map_unit_button (driver)
        time.sleep (5)
        
        locate_and_click_view_rating_for_hydric_rating_by_map_unit_button(driver)
        time.sleep (5)
        
        #scroll_to_bottom_of_the_page (driver)

        data_in_hydric_rating_by_map_unit = extract_table_data_two (driver)
        print (data_in_hydric_rating_by_map_unit)
        
        map_unit_symbol, map_unit_name, rating = find_highest_percent_aoi (data_in_hydric_rating_by_map_unit)
        
        print ('map_unit_symbol :', map_unit_symbol , 'map_unit_name ', map_unit_name, 'rating', rating)
        
        time.sleep (10)


except Exception as e:
    print("Error occurred: ", e)
finally:
    driver.quit()

